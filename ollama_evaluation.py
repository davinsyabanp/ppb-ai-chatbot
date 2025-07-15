import pandas as pd
import numpy as np
from app.core import get_response_for_evaluation
import os
from dotenv import load_dotenv
import time
import re
import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Load environment variables
load_dotenv()

class OllamaEvaluator:
    """
    Local LLM evaluator using Ollama.
    """
    def __init__(self, model_name="llama3.2:latest", base_url="http://localhost:11434"):
        self.model_name = model_name
        self.base_url = base_url
        self.api_url = f"{base_url}/api/generate"
        
        # Test connection
        self._test_connection()
    
    def _test_connection(self):
        """Test if Ollama is running and the model is available."""
        try:
            # Check if Ollama is running
            response = requests.get(f"{self.base_url}/api/tags")
            if response.status_code != 200:
                raise Exception("Ollama is not running. Please start Ollama first.")
            
            # Check if model exists
            models = response.json().get("models", [])
            model_names = [model["name"] for model in models]
            
            if self.model_name not in model_names:
                print(f"Model {self.model_name} not found. Available models: {model_names}")
                print("You can pull the model using: ollama pull llama3.2:3b")
                raise Exception(f"Model {self.model_name} not available")
            
            print(f"✅ Connected to Ollama with model: {self.model_name}")
            
        except requests.exceptions.ConnectionError:
            raise Exception("Cannot connect to Ollama. Please make sure Ollama is running on http://localhost:11434")
        except Exception as e:
            raise Exception(f"Error testing Ollama connection: {e}")
    
    def call_ollama(self, prompt, max_retries=3):
        """
        Call Ollama with a prompt and return the response.
        """
        payload = {
            "model": self.model_name,
            "prompt": prompt,
            "stream": False,
            "options": {
                "temperature": 0.1,
                "top_p": 0.9,
                "max_tokens": 500
            }
        }
        
        for attempt in range(max_retries):
            try:
                response = requests.post(self.api_url, json=payload, timeout=300)
                if response.status_code == 200:
                    return response.json()["response"]
                else:
                    print(f"Attempt {attempt + 1}: HTTP {response.status_code}")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)  # Exponential backoff
            except requests.exceptions.RequestException as e:
                print(f"Attempt {attempt + 1}: Request failed - {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
        
        return "Error: Failed to get response from Ollama"
    
    def extract_score_from_response(self, response):
        """
        Extract numerical score from Ollama response.
        Looks for patterns like 'Skor: 8/10' or '8/10' or '8'.
        """
        # Look for patterns like "Skor: 8/10", "8/10", "8", etc.
        patterns = [
            r'Skor:\s*(\d+)/10',
            r'(\d+)/10',
            r'Skor:\s*(\d+)',
            r'(\d+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, response)
            if match:
                score = int(match.group(1))
                # Ensure score is between 0 and 10
                return max(0, min(10, score))
        
        # If no score found, try to infer from text
        response_lower = response.lower()
        if 'sangat baik' in response_lower or 'excellent' in response_lower:
            return 9
        elif 'baik' in response_lower or 'good' in response_lower:
            return 7
        elif 'cukup' in response_lower or 'fair' in response_lower:
            return 5
        elif 'kurang' in response_lower or 'poor' in response_lower:
            return 3
        else:
            return 5  # Default score

def evaluate_faithfulness(evaluator, context, answer):
    """
    Evaluate if the answer faithfully reflects the information in the context.
    """
    prompt = f"""
    Tugas: Evaluasi kesetiaan jawaban terhadap konteks yang diberikan.
    
    Konteks:
    {context}
    
    Jawaban:
    {answer}
    
    Pertanyaan: Apakah jawaban tersebut setia dan akurat mencerminkan informasi yang ada dalam konteks? 
    
    Berikan penilaian dengan skala 1-10, di mana:
    10 = Sangat setia, semua informasi akurat
    5 = Cukup setia, sebagian besar informasi akurat
    1 = Tidak setia, banyak informasi tidak akurat atau tidak ada dalam konteks
    
    Berikan juga penjelasan singkat mengapa Anda memberikan skor tersebut.
    
    Format jawaban:
    Penjelasan: [penjelasan Anda]
    Skor: [angka]/10
    """
    
    response = evaluator.call_ollama(prompt)
    score = evaluator.extract_score_from_response(response)
    
    return {
        'score': score,
        'explanation': response,
        'metric': 'faithfulness'
    }

def evaluate_answer_relevancy(evaluator, question, answer):
    """
    Evaluate if the answer is relevant and directly addresses the question.
    """
    prompt = f"""
    Tugas: Evaluasi relevansi jawaban terhadap pertanyaan.
    
    Pertanyaan: {question}
    Jawaban: {answer}
    
    Pertanyaan: Apakah jawaban tersebut relevan dan langsung menjawab pertanyaan yang diajukan?
    
    Berikan penilaian dengan skala 1-10, di mana:
    10 = Sangat relevan, langsung menjawab pertanyaan
    5 = Cukup relevan, sebagian menjawab pertanyaan
    1 = Tidak relevan, tidak menjawab pertanyaan
    
    Berikan juga penjelasan singkat mengapa Anda memberikan skor tersebut.
    
    Format jawaban:
    Penjelasan: [penjelasan Anda]
    Skor: [angka]/10
    """
    
    response = evaluator.call_ollama(prompt)
    score = evaluator.extract_score_from_response(response)
    
    return {
        'score': score,
        'explanation': response,
        'metric': 'answer_relevancy'
    }

def evaluate_context_precision(evaluator, question, contexts):
    """
    Evaluate if the retrieved contexts are relevant to the question.
    """
    context_text = "\n\n".join(contexts) if contexts else "Tidak ada konteks"
    
    prompt = f"""
    Tugas: Evaluasi presisi konteks yang diambil untuk pertanyaan.
    
    Pertanyaan: {question}
    
    Konteks yang diambil:
    {context_text}
    
    Pertanyaan: Apakah konteks yang diambil relevan dan tepat untuk menjawab pertanyaan tersebut?
    
    Berikan penilaian dengan skala 1-10, di mana:
    10 = Sangat presisi, semua konteks sangat relevan
    5 = Cukup presisi, sebagian besar konteks relevan
    1 = Tidak presisi, sebagian besar konteks tidak relevan
    
    Berikan juga penjelasan singkat mengapa Anda memberikan skor tersebut.
    
    Format jawaban:
    Penjelasan: [penjelasan Anda]
    Skor: [angka]/10
    """
    
    response = evaluator.call_ollama(prompt)
    score = evaluator.extract_score_from_response(response)
    
    return {
        'score': score,
        'explanation': response,
        'metric': 'context_precision'
    }

def evaluate_context_recall(evaluator, question, ground_truth, contexts):
    """
    Evaluate if the retrieved contexts contain the necessary information to answer the question.
    """
    context_text = "\n\n".join(contexts) if contexts else "Tidak ada konteks"
    
    prompt = f"""
    Tugas: Evaluasi recall konteks - apakah konteks yang diambil mengandung informasi yang diperlukan.
    
    Pertanyaan: {question}
    
    Jawaban yang diharapkan (ground truth):
    {ground_truth}
    
    Konteks yang diambil:
    {context_text}
    
    Pertanyaan: Apakah konteks yang diambil mengandung informasi yang diperlukan untuk memberikan jawaban yang diharapkan?
    
    Berikan penilaian dengan skala 1-10, di mana:
    10 = Sangat lengkap, semua informasi yang diperlukan ada
    5 = Cukup lengkap, sebagian besar informasi ada
    1 = Tidak lengkap, banyak informasi yang diperlukan tidak ada
    
    Berikan juga penjelasan singkat mengapa Anda memberikan skor tersebut.
    
    Format jawaban:
    Penjelasan: [penjelasan Anda]
    Skor: [angka]/10
    """
    
    response = evaluator.call_ollama(prompt)
    score = evaluator.extract_score_from_response(response)
    
    return {
        'score': score,
        'explanation': response,
        'metric': 'context_recall'
    }

def save_excel_results(results_df, avg_scores, model_name):
    """
    Save evaluation results in Excel format with proper formatting.
    """
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Summary sheet
        summary_sheet = wb.create_sheet("Summary")
        
        # Add title and metadata
        summary_sheet['A1'] = "OLLAMA RAG EVALUATION SUMMARY"
        summary_sheet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        summary_sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        summary_sheet.merge_cells('A1:C1')
        
        # Add metadata
        summary_sheet['A3'] = "Evaluation Date:"
        summary_sheet['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        summary_sheet['A4'] = "Model Used:"
        summary_sheet['B4'] = model_name
        summary_sheet['A5'] = "Total Questions:"
        summary_sheet['B5'] = len(results_df)
        
        # Add metric scores
        summary_sheet['A7'] = "METRIC"
        summary_sheet['B7'] = "SCORE"
        summary_sheet['C7'] = "GRADE"
        
        # Style headers
        for cell in ['A7', 'B7', 'C7']:
            summary_sheet[cell].font = Font(bold=True)
            summary_sheet[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Add metrics
        row = 8
        for metric, score in avg_scores.items():
            summary_sheet[f'A{row}'] = metric.replace('_', ' ').title()
            summary_sheet[f'B{row}'] = round(score, 2)
            
            # Add grade based on score
            if score >= 9:
                grade = "Excellent"
                color = "C6EFCE"  # Green
            elif score >= 7:
                grade = "Good"
                color = "FFEB9C"  # Yellow
            elif score >= 5:
                grade = "Fair"
                color = "FFC7CE"  # Light red
            else:
                grade = "Poor"
                color = "FFC7CE"  # Red
            
            summary_sheet[f'C{row}'] = grade
            summary_sheet[f'C{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            row += 1
        
        # Add interpretation guide
        summary_sheet['A12'] = "INTERPRETATION GUIDE"
        summary_sheet['A12'].font = Font(bold=True, size=12)
        summary_sheet['A13'] = "9-10: Excellent"
        summary_sheet['A14'] = "7-8:  Good"
        summary_sheet['A15'] = "5-6:  Fair"
        summary_sheet['A16'] = "3-4:  Poor"
        summary_sheet['A17'] = "1-2:  Very Poor"
        
        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create Detailed Results sheet
        details_sheet = wb.create_sheet("Detailed Results")
        
        # Add title
        details_sheet['A1'] = "DETAILED EVALUATION RESULTS"
        details_sheet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        details_sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        details_sheet.merge_cells('A1:H1')
        
        # Add headers
        headers = [
            'Question', 'Ground Truth', 'Answer', 'Contexts Count',
            'Faithfulness Score', 'Answer Relevancy Score', 
            'Context Precision Score', 'Context Recall Score'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = details_sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for row_idx, (_, row_data) in enumerate(results_df.iterrows(), 4):
            details_sheet.cell(row=row_idx, column=1, value=row_data['question'])
            details_sheet.cell(row=row_idx, column=2, value=row_data['ground_truth'])
            details_sheet.cell(row=row_idx, column=3, value=row_data['answer'])
            details_sheet.cell(row=row_idx, column=4, value=row_data['contexts_count'])
            details_sheet.cell(row=row_idx, column=5, value=row_data['faithfulness_score'])
            details_sheet.cell(row=row_idx, column=6, value=row_data['relevancy_score'])
            details_sheet.cell(row=row_idx, column=7, value=row_data['precision_score'])
            details_sheet.cell(row=row_idx, column=8, value=row_data['recall_score'])
        
        # Auto-adjust column widths for details sheet
        for column in details_sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 80)
                details_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create Explanations sheet
        explanations_sheet = wb.create_sheet("Explanations")
        
        # Add title
        explanations_sheet['A1'] = "DETAILED EXPLANATIONS"
        explanations_sheet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        explanations_sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        explanations_sheet.merge_cells('A1:E1')
        
        # Add headers for explanations
        exp_headers = [
            'Question', 'Faithfulness Explanation', 'Relevancy Explanation',
            'Precision Explanation', 'Recall Explanation'
        ]
        
        for col, header in enumerate(exp_headers, 1):
            cell = explanations_sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add explanation data
        for row_idx, (_, row_data) in enumerate(results_df.iterrows(), 4):
            explanations_sheet.cell(row=row_idx, column=1, value=row_data['question'])
            explanations_sheet.cell(row=row_idx, column=2, value=row_data['faithfulness_explanation'])
            explanations_sheet.cell(row=row_idx, column=3, value=row_data['relevancy_explanation'])
            explanations_sheet.cell(row=row_idx, column=4, value=row_data['precision_explanation'])
            explanations_sheet.cell(row=row_idx, column=5, value=row_data['recall_explanation'])
        
        # Auto-adjust column widths for explanations sheet
        for column in explanations_sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 100)
                explanations_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        excel_filename = f'ollama_evaluation_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(excel_filename)
        
        print(f"📊 Excel results saved to: {excel_filename}")
        
        return excel_filename
        
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")
        print("📄 CSV files were still saved successfully")
        return None

def run_ollama_evaluation(csv_path="evaluation/eval_dataset_PPB.csv", sample_size=None, model_name="llama3.2:latest"):
    """
    Run the complete Ollama-based evaluation.
    """
    print(f"Starting evaluation with model: {model_name}")
    print(f"CSV path: {csv_path}")
    
    if not os.path.exists(csv_path):
        print(f"❌ Error: CSV file not found at {csv_path}")
        print(f"Current working directory: {os.getcwd()}")
        print(f"Available files in current directory: {os.listdir('.')}")
        return None, None
    
    # Initialize Ollama evaluator
    try:
        print("🔧 Initializing Ollama evaluator...")
        evaluator = OllamaEvaluator(model_name=model_name)
        print("✅ Ollama evaluator initialized successfully")
    except Exception as e:
        print(f"❌ Error initializing Ollama: {e}")
        print("Please make sure:")
        print("1. Ollama is installed and running (ollama serve)")
        print("2. The model is available (ollama pull llama3.2:latest)")
        print("3. Ollama is accessible at http://localhost:11434")
        return None, None
    
    # Load dataset
    try:
        print("📊 Loading evaluation dataset...")
        df = pd.read_csv(csv_path)
        print(f"✅ Loaded {len(df)} questions from {csv_path}")
    except Exception as e:
        print(f"❌ Error loading CSV: {e}")
        return None, None
    
    # Determine evaluation size
    if sample_size is None:
        # Evaluate all questions
        evaluation_df = df
        print(f"🎯 Evaluating ALL {len(evaluation_df)} questions using {model_name}...")
        print("⚠️  This may take a while depending on the dataset size and model speed.")
    else:
        # Take a sample
        evaluation_df = df.head(sample_size)
        print(f"🎯 Evaluating {len(evaluation_df)} questions (sample) using {model_name}...")
    
    results = []
    
    for idx, row in evaluation_df.iterrows():
        question = row['question']
        ground_truth = row['ground_truth']
        
        print(f"\n--- Processing Question {idx + 1}/{len(evaluation_df)} ---")
        print(f"Question: {question}")
        
        # Get RAG response
        try:
            print("🔄 Getting RAG response...")
            rag_result = get_response_for_evaluation(question, user_id="evaluation_user")
            answer = rag_result['answer']
            contexts = rag_result['contexts']
            
            print(f"Answer: {answer[:100]}...")
            print(f"Number of contexts: {len(contexts)}")
        except Exception as e:
            print(f"❌ Error getting RAG response: {e}")
            continue
        
        # Evaluate each metric
        try:
            print("📝 Evaluating faithfulness...")
            faithfulness_result = evaluate_faithfulness(evaluator, "\n\n".join(contexts), answer)
            
            print("📝 Evaluating answer relevancy...")
            relevancy_result = evaluate_answer_relevancy(evaluator, question, answer)
            
            print("📝 Evaluating context precision...")
            precision_result = evaluate_context_precision(evaluator, question, contexts)
            
            print("📝 Evaluating context recall...")
            recall_result = evaluate_context_recall(evaluator, question, ground_truth, contexts)
        except Exception as e:
            print(f"❌ Error during evaluation: {e}")
            continue
        
        # Store results
        result = {
            'question': question,
            'ground_truth': ground_truth,
            'answer': answer,
            'contexts_count': len(contexts),
            'faithfulness_score': faithfulness_result['score'],
            'faithfulness_explanation': faithfulness_result['explanation'],
            'relevancy_score': relevancy_result['score'],
            'relevancy_explanation': relevancy_result['explanation'],
            'precision_score': precision_result['score'],
            'precision_explanation': precision_result['explanation'],
            'recall_score': recall_result['score'],
            'recall_explanation': recall_result['explanation']
        }
        
        results.append(result)
        print(f"✅ Question {idx + 1} completed successfully")
        
        # Add delay to avoid overwhelming local resources
        time.sleep(1)
        
        # Show progress every 10 questions
        if (idx + 1) % 10 == 0:
            print(f"📊 Progress: {idx + 1}/{len(evaluation_df)} questions completed ({((idx + 1)/len(evaluation_df)*100):.1f}%)")
    
    if not results:
        print("❌ No results were generated. Evaluation failed.")
        return None, None
    
    # Create results DataFrame
    results_df = pd.DataFrame(results)
    
    # Calculate average scores
    avg_scores = {
        'faithfulness': results_df['faithfulness_score'].mean(),
        'answer_relevancy': results_df['relevancy_score'].mean(),
        'context_precision': results_df['precision_score'].mean(),
        'context_recall': results_df['recall_score'].mean()
    }
    
    # Save detailed results
    results_df.to_csv('ollama_evaluation_results.csv', index=False)
    
    # Save summary
    summary_df = pd.DataFrame([avg_scores])
    summary_df.to_csv('ollama_evaluation_summary.csv', index=False)
    
    # Save Excel format with formatting
    save_excel_results(results_df, avg_scores, model_name)
    
    print(f"✅ Evaluation completed successfully!")
    print(f"📊 Processed {len(results)} questions")
    
    return results_df, avg_scores

def print_summary(avg_scores):
    """
    Print a formatted summary of the evaluation results.
    """
    print("\n" + "="*60)
    print("OLLAMA RAG EVALUATION SUMMARY")
    print("="*60)
    
    for metric, score in avg_scores.items():
        print(f"{metric.replace('_', ' ').title()}: {score:.2f}/10")
    
    print("\n" + "="*60)
    print("INTERPRETATION:")
    print("="*60)
    print("9-10: Excellent")
    print("7-8:  Good")
    print("5-6:  Fair")
    print("3-4:  Poor")
    print("1-2:  Very Poor")
    print("="*60)

def check_ollama_models():
    """
    Check available Ollama models and suggest the best one for evaluation.
    """
    try:
        response = requests.get("http://localhost:11434/api/tags")
        if response.status_code == 200:
            models = response.json().get("models", [])
            print("Available Ollama models:")
            for model in models:
                print(f"  - {model['name']} ({model.get('size', 'Unknown size')})")
            
            # Suggest best models for evaluation
            print("\nRecommended models for evaluation (in order of preference):")
            recommended = [
                "llama3.2:3b",
                "llama3.1:8b", 
                "llama3.1:3b",
                "mistral:7b",
                "qwen2.5:3b"
            ]
            
            for model in recommended:
                if any(model in m['name'] for m in models):
                    print(f"  ✅ {model} - Available")
                else:
                    print(f"  ❌ {model} - Not available (run: ollama pull {model})")
            
            return models
        else:
            print("Cannot connect to Ollama. Make sure it's running.")
            return []
    except Exception as e:
        print(f"Error checking Ollama models: {e}")
        return []

def get_evaluation_mode():
    """
    Let user choose between evaluating all questions or a sample.
    """
    print("\n" + "="*60)
    print("EVALUATION MODE SELECTION")
    print("="*60)
    print("Choose your evaluation mode:")
    print("1. Evaluate ALL questions (comprehensive but takes longer)")
    print("2. Evaluate a SAMPLE of questions (faster for testing)")
    print("3. Cancel evaluation")
    
    while True:
        try:
            choice = input("\nEnter your choice (1, 2, or 3): ").strip()
            
            if choice == "1":
                return None  # None means evaluate all
            elif choice == "2":
                while True:
                    try:
                        sample_size = input("Enter sample size (e.g., 5, 10, 20): ").strip()
                        sample_size = int(sample_size)
                        if sample_size > 0:
                            return sample_size
                        else:
                            print("❌ Sample size must be greater than 0")
                    except ValueError:
                        print("❌ Please enter a valid number")
            elif choice == "3":
                print("Evaluation cancelled.")
                return "cancel"
            else:
                print("❌ Please enter 1, 2, or 3")
        except KeyboardInterrupt:
            print("\nEvaluation cancelled.")
            return "cancel"

if __name__ == "__main__":
    print("Starting Ollama-based RAG Evaluation...")
    print("This will evaluate your RAG system using a local LLM with Indonesian prompts.")
    
    # Check available models
    print("\nChecking available Ollama models...")
    available_models = check_ollama_models()
    
    if not available_models:
        print("\n❌ No Ollama models found. Please:")
        print("1. Install Ollama: https://ollama.ai/")
        print("2. Start Ollama: ollama serve")
        print("3. Pull a model: ollama pull llama3.2:3b")
        exit(1)
    
    # Choose model (you can modify this)
    model_name = "deepseek-r1:8b"  # Change this to your preferred model
    
    print(f"\nUsing model: {model_name}")
    print("You can change the model by modifying the 'model_name' variable in the script.")
    
    # Get evaluation mode from user
    sample_size = get_evaluation_mode()
    
    if sample_size == "cancel":
        exit(0)
    
    # Run evaluation
    results, avg_scores = run_ollama_evaluation(sample_size=sample_size, model_name=model_name)
    
    if results is not None and avg_scores is not None:
        # Print summary
        print_summary(avg_scores)
        
        print(f"\n📄 CSV files saved:")
        print(f"  - ollama_evaluation_results.csv")
        print(f"  - ollama_evaluation_summary.csv")
        print(f"📊 Excel file saved with detailed formatting and multiple sheets")
        
        # Show first few results
        print(f"\nFirst 3 results preview:")
        print(results[['question', 'faithfulness_score', 'relevancy_score', 'precision_score', 'recall_score']].head(3))
    else:
        print("❌ Evaluation failed. Please check the error messages above.")
        print("\n🔧 Troubleshooting steps:")
        print("1. Make sure Ollama is running: ollama serve")
        print("2. Check if the model is available: ollama list")
        print("3. Pull the model if needed: ollama pull llama3.2:latest")
        print("4. Check if the CSV file exists: evaluation/eval_dataset_PPB.csv") 